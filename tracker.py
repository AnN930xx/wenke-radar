"""投递记录层：applications.db 的唯一入口 + Excel 导出。

投递记录是人工数据（不可再生），单独一个库并**留在 git** 持久化；
与可再生的抓取库 jobs.db（不进 git，云端 Actions cache 接力）严格分开。
report 层要投递数据一律经 get_recent_applications() 拿，不许直连本库。

命令行:
    python tracker.py add          # 交互式登记一条投递
    python tracker.py update       # 更新状态/反馈/下一步
    python tracker.py list [状态]  # 列出（可按状态过滤）
    python tracker.py stats        # 统计
    python tracker.py export       # 导出带配色的 Excel
"""
import os
import sqlite3
import sys
from contextlib import closing
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ROOT = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(_ROOT, "data", "applications.db")
EXPORT_PATH = os.path.join(_ROOT, "reports", "投递追踪表.xlsx")

# 状态流转顺序 + Excel 配色（十六进制，不带 #）
STATUS_FLOW = [
    ("待投递", "9E9E9E"),   # 灰：收藏未投
    ("已投递", "2196F3"),   # 蓝
    ("笔试",   "FF9800"),   # 橙
    ("一面",   "9C27B0"),   # 紫：业务面
    ("二面",   "9C27B0"),
    ("三面",   "9C27B0"),
    ("HR面",   "00BCD4"),   # 青：快出结果
    ("Offer",  "4CAF50"),   # 绿 🎉
    ("已拒",   "F44336"),   # 红
    ("已放弃", "795548"),   # 棕
]
STATUS_NAMES = [name for name, _ in STATUS_FLOW]
STATUS_COLORS = dict(STATUS_FLOW)

CHANNELS = ["官网", "内推", "牛客", "邮件", "招聘会", "其他"]

_SCHEMA = """
    CREATE TABLE IF NOT EXISTS applications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company TEXT NOT NULL,         -- 公司名称
        position TEXT NOT NULL,        -- 岗位名称
        apply_date TEXT,               -- 投递日期
        channel TEXT,                  -- 投递渠道
        status TEXT DEFAULT '已投递',   -- 当前状态
        feedback TEXT,                 -- 反馈内容
        next_step TEXT,                -- 下一步动作
        job_url TEXT,                  -- 岗位链接
        update_time TEXT,              -- 最后更新时间
        remark TEXT                    -- 备注
    )
"""


def _connect():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute(_SCHEMA)
    return conn


def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M")


# ==================== 读写操作 ====================

def add_record(company, position, channel="官网", status="已投递",
               feedback="", next_step="", job_url="", remark=""):
    """登记一条投递"""
    with closing(_connect()) as conn:
        conn.execute(
            "INSERT INTO applications (company, position, apply_date, channel, status,"
            " feedback, next_step, job_url, update_time, remark)"
            " VALUES (?,?,?,?,?,?,?,?,?,?)",
            (company, position, date.today().isoformat(), channel, status,
             feedback, next_step, job_url, _now(), remark))
        conn.commit()
    print(f"  ✅ 已添加: {company} - {position} [{status}]")


def update_record(record_id, status=None, feedback=None, next_step=None, remark=None):
    """更新一条投递的可变字段（None=保持不变）"""
    changes = {"update_time": _now()}
    if status:
        changes["status"] = status
    if feedback is not None:
        changes["feedback"] = feedback
    if next_step is not None:
        changes["next_step"] = next_step
    if remark is not None:
        changes["remark"] = remark
    assignments = ", ".join(f"{col} = ?" for col in changes)
    with closing(_connect()) as conn:
        conn.execute(f"UPDATE applications SET {assignments} WHERE id = ?",
                     (*changes.values(), record_id))
        conn.commit()
    print(f"  ✅ 已更新记录 #{record_id}")


def get_recent_applications():
    """report 层的唯一取数口。返回按更新时间倒序的元组列表：
    (company, position, status, apply_date, feedback, next_step)"""
    with closing(_connect()) as conn:
        return conn.execute(
            "SELECT company, position, status, apply_date, feedback, next_step"
            " FROM applications ORDER BY update_time DESC").fetchall()


def list_records(status_filter=None):
    """列出投递记录（可按状态过滤），并打印"""
    sql = "SELECT * FROM applications"
    args = ()
    if status_filter:
        sql += " WHERE status = ?"
        args = (status_filter,)
    sql += " ORDER BY apply_date DESC"
    with closing(_connect()) as conn:
        rows = conn.execute(sql, args).fetchall()
    _print_records(rows)
    return rows


def get_stats():
    """终端投递统计"""
    with closing(_connect()) as conn:
        total = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
        by_status = conn.execute(
            "SELECT status, COUNT(*) FROM applications GROUP BY status"
            " ORDER BY COUNT(*) DESC").fetchall()
        by_company = conn.execute(
            "SELECT company, COUNT(*) FROM applications GROUP BY company"
            " ORDER BY COUNT(*) DESC LIMIT 10").fetchall()

    print(f"\n{'=' * 50}")
    print(f"📊 投递统计 ｜ 共投递 {total} 个岗位")
    print(f"{'=' * 50}")
    print("\n按状态分布:")
    for status, n in by_status:
        print(f"  {status:6s} {n:3d} {'█' * n}")
    if by_company:
        print("\n投递最多的公司:")
        for company, n in by_company:
            print(f"  {company:12s} {n}")
    print()
    return {"total": total, "by_status": dict(by_status)}


# ==================== Excel 导出 ====================

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def export_excel():
    """导出两个 sheet：投递总表（状态列配色）+ 统计看板"""
    with closing(_connect()) as conn:
        rows = conn.execute(
            "SELECT id, company, position, apply_date, channel, status,"
            " feedback, next_step, job_url, update_time, remark"
            " FROM applications ORDER BY apply_date DESC, update_time DESC").fetchall()
        total = len(rows)
        by_status = conn.execute(
            "SELECT status, COUNT(*) FROM applications GROUP BY status").fetchall()
        by_channel = conn.execute(
            "SELECT channel, COUNT(*) FROM applications GROUP BY channel").fetchall()
        by_company = conn.execute(
            "SELECT company, COUNT(*) FROM applications GROUP BY company"
            " ORDER BY COUNT(*) DESC").fetchall()

    wb = Workbook()
    _build_main_sheet(wb.active, rows)
    _build_stats_sheet(wb.create_sheet("统计看板"), total, by_status, by_channel, by_company)

    os.makedirs(os.path.dirname(EXPORT_PATH), exist_ok=True)
    wb.save(EXPORT_PATH)
    print(f"\n  📊 Excel 已导出: {EXPORT_PATH}")
    print(f"  共 {len(rows)} 条投递记录")
    return EXPORT_PATH


def _build_main_sheet(ws, rows):
    ws.title = "投递总表"
    headers = ["序号", "公司", "岗位", "投递日期", "渠道",
               "状态", "反馈", "下一步", "岗位链接", "更新时间", "备注"]
    STATUS_COL = 6

    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    for r, record in enumerate(rows, 2):
        for c, value in enumerate(record, 1):
            cell = ws.cell(row=r, column=c, value=value or "")
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == STATUS_COL and value:
                color = STATUS_COLORS.get(value, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

    for i, width in enumerate([6, 14, 28, 12, 8, 10, 35, 20, 30, 18, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _build_stats_sheet(ws, total, by_status, by_channel, by_company):
    def title_cell(row, text, size=12):
        ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=size)

    def block(start_row, heading, col_names, data):
        """一段 标题+两列表格，返回下一可用行"""
        title_cell(start_row, heading)
        for c, name in enumerate(col_names, 1):
            ws.cell(row=start_row + 1, column=c, value=name).font = Font(bold=True)
        for i, (key, n) in enumerate(data, start_row + 2):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=n)
            if len(col_names) > 2:
                ws.cell(row=i, column=3,
                        value=f"{n / total * 100:.1f}%" if total else "0%")
        return start_row + 2 + len(data)

    title_cell(1, "投递统计看板", size=14)
    ws.cell(row=2, column=1, value=f"更新时间: {_now()}")
    title_cell(4, f"总投递数: {total}")

    row = block(6, "按状态分布", ["状态", "数量", "占比"], by_status)
    row = block(row + 2, "按渠道分布", ["渠道", "数量"], by_channel)
    block(row + 2, "投递公司一览", ["公司", "投递数"], by_company)

    for col, width in (("A", 20), ("B", 12), ("C", 12)):
        ws.column_dimensions[col].width = width


# ==================== 终端打印与交互 ====================

def _print_records(rows):
    if not rows:
        print("\n  暂无投递记录。用 'python tracker.py add' 添加第一条吧。")
        return
    print(f"\n{'=' * 100}")
    print(f"投递记录 ｜ 共 {len(rows)} 条")
    print(f"{'=' * 100}")
    for r in rows:
        rid, company, position = r[0], r[1], r[2]
        apply_date, channel, status = r[3], r[4], r[5]
        feedback, next_step, updated, remark = r[6], r[7], r[9], r[10]
        print(f"\n  #{rid} [{status}] {company} - {position}")
        print(f"     投递: {apply_date}  渠道: {channel}  更新: {updated}")
        for label, text in (("反馈", feedback), ("下一步", next_step), ("备注", remark)):
            if text:
                print(f"     {label}: {text}")


def _ask(prompt, default=""):
    value = input(prompt).strip()
    return value or default


def interactive_add():
    print("\n📝 添加投递记录")
    print("-" * 30)
    company = _ask("公司名称: ")
    position = _ask("岗位名称: ")
    if not company or not position:
        print("公司名和岗位名都不能为空")
        return
    channel = _ask(f"投递渠道 ({'/'.join(CHANNELS)}) [默认: 官网]: ", "官网")
    status = _ask(f"当前状态 ({'/'.join(STATUS_NAMES)}) [默认: 已投递]: ", "已投递")
    if status not in STATUS_NAMES:
        print(f"  未知状态 '{status}'，已设为'已投递'")
        status = "已投递"
    feedback = _ask("反馈（收到笔试/面试通知等，可留空）: ")
    next_step = _ask("下一步（准备笔试/等通知等，可留空）: ")
    remark = _ask("备注（可留空）: ")
    add_record(company, position, channel, status, feedback, next_step, "", remark)


def interactive_update():
    if not list_records():
        return
    rid = _ask("\n输入要更新的记录序号 (#): ")
    if not rid.isdigit():
        print("序号必须是数字")
        return
    with closing(_connect()) as conn:
        row = conn.execute("SELECT * FROM applications WHERE id = ?",
                           (int(rid),)).fetchone()
    if not row:
        print(f"找不到记录 #{rid}")
        return

    print(f"\n当前记录: #{row[0]} [{row[5]}] {row[1]} - {row[2]}")
    print(f"当前反馈: {row[6] or '无'}")
    print(f"当前下一步: {row[7] or '无'}\n")
    status = _ask(f"新状态 ({'/'.join(STATUS_NAMES)}) [回车保持不变]: ") or None
    feedback = _ask("新反馈 [回车保持不变]: ") or None
    next_step = _ask("新下一步 [回车保持不变]: ") or None
    remark = _ask("新备注 [回车保持不变]: ") or None
    update_record(int(rid), status, feedback, next_step, remark)


def interactive_main():
    if len(sys.argv) < 2:
        print(__doc__)
        return
    command = sys.argv[1]
    actions = {
        "add": interactive_add,
        "update": interactive_update,
        "list": lambda: list_records(sys.argv[2] if len(sys.argv) > 2 else None),
        "export": export_excel,
        "stats": get_stats,
    }
    if command in actions:
        actions[command]()
    else:
        print(f"未知命令: {command}")
        print(__doc__)


if __name__ == "__main__":
    interactive_main()
